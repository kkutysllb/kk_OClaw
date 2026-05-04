"""
XLSX Creator Skill — generate.py

Create and manipulate Excel files using openpyxl:
  - create    Create a new workbook from CSV, JSON, or inline data
  - add-chart Insert a chart (bar, line, pie, scatter) into an existing sheet
  - format    Apply cell formatting (fonts, colors, borders, alignment)
  - info      Inspect workbook structure (sheets, columns, row counts)
"""

import json
import os
import sys

try:
    import openpyxl
    from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    print("⚠  openpyxl is not installed. Run: pip install openpyxl")
    sys.exit(1)


# ── Helpers ──────────────────────────────────────────────────────────
def _auto_width(ws) -> None:
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)


def _load_data(source: str) -> tuple[list[str], list[list]]:
    """Load tabular data from a CSV or JSON file.

    Returns (headers, rows).
    CSV: first row = headers.
    JSON: array of objects → keys are headers; or {headers: [...], rows: [[...]]}.
    """
    ext = os.path.splitext(source)[1].lower()

    if ext == ".csv":
        import csv
        with open(source, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)
        if not rows:
            return [], []
        return rows[0], rows[1:]

    if ext == ".json":
        with open(source, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list) and data:
            headers = list(data[0].keys())
            rows = [[row.get(h, "") for h in headers] for row in data]
            return headers, rows
        if isinstance(data, dict):
            # Support {headers: [...], rows: [[...]]}
            headers = data.get("headers", data.get("columns", []))
            rows = data.get("rows", data.get("data", []))
            return headers, rows
        return [], []

    return [], []


# ── create ───────────────────────────────────────────────────────────
def create_workbook(
    source: str,
    output_path: str,
    sheet_name: str = "Sheet1",
    auto_width: bool = True,
    header_style: bool = True,
) -> str:
    """Create a new Excel workbook from a data source.

    Args:
        source: Path to CSV or JSON data file.
        output_path: Output .xlsx path.
        sheet_name: Name for the worksheet.
        auto_width: Auto-fit column widths.
        header_style: Apply bold + colored header row.
    """
    headers, rows = _load_data(source)
    if not headers:
        return "Error: could not parse data from source file."

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Write headers
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        if header_style:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

    # Write data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, val in enumerate(row_data, 1):
            # Try to convert numeric strings
            if isinstance(val, str):
                try:
                    val = int(val)
                except ValueError:
                    try:
                        val = float(val)
                    except ValueError:
                        pass
            ws.cell(row=row_idx, column=col_idx, value=val)

    if auto_width:
        _auto_width(ws)

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    return (
        f"Created workbook '{sheet_name}' with {len(headers)} column(s) "
        f"and {len(rows)} data row(s) → {output_path}"
    )


# ── add-chart ────────────────────────────────────────────────────────
def add_chart(
    workbook_path: str,
    sheet_name: str,
    chart_type: str,
    data_start: str,
    data_end: str,
    categories_col: int = 1,
    output_path: str = "",
    title: str = "",
    x_axis: str = "",
    y_axis: str = "",
) -> str:
    """Add a chart to an existing workbook.

    Args:
        workbook_path: Path to the .xlsx file.
        sheet_name: Name of the sheet containing the data.
        chart_type: One of bar, line, pie, scatter.
        data_start: Top-left cell of data range (e.g. "B1").
        data_end: Bottom-right cell of data range (e.g. "D10").
        categories_col: Column index (1-based) for category labels.
        output_path: Output path (defaults to overwriting input).
        title: Chart title.
        x_axis: X-axis label.
        y_axis: Y-axis label.
    """
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb[sheet_name]

    # Parse data range
    start_col_letter = "".join(c for c in data_start if c.isalpha())
    start_row = int("".join(c for c in data_start if c.isdigit()))
    end_col_letter = "".join(c for c in data_end if c.isalpha())
    end_row = int("".join(c for c in data_end if c.isdigit()))

    # Categories (first column of data range by default)
    cats = Reference(ws, min_col=categories_col, min_row=start_row,
                     max_row=end_row)

    # Data series (remaining columns)
    start_col_idx = openpyxl.utils.column_index_from_string(start_col_letter)
    end_col_idx = openpyxl.utils.column_index_from_string(end_col_letter)

    # Build chart
    chart_map = {
        "bar": BarChart,
        "line": LineChart,
        "pie": PieChart,
        "scatter": ScatterChart,
    }
    chart_cls = chart_map.get(chart_type.lower())
    if not chart_cls:
        return f"Error: unknown chart type '{chart_type}'. Use bar, line, pie, or scatter."

    chart = chart_cls()
    if title:
        chart.title = title
    if x_axis:
        chart.x_axis.title = x_axis
    if y_axis:
        chart.y_axis.title = y_axis

    if chart_type == "pie":
        data = Reference(ws, min_col=start_col_idx, min_row=start_row,
                         max_row=end_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
    else:
        for col_idx in range(start_col_idx, end_col_idx + 1):
            if col_idx == categories_col:
                continue
            col_letter = get_column_letter(col_idx)
            data = Reference(ws, min_col=col_idx, min_row=start_row - 1 if start_row > 1 else 1,
                             max_row=end_row)
            chart.add_data(data, titles_from_data=(start_row == 1))
        if start_row == 1:
            chart.set_categories(cats)
        else:
            chart.set_categories(Reference(ws, min_col=categories_col, min_row=start_row,
                                           max_row=end_row))

    # Place chart below data
    chart_row = end_row + 3
    ws.add_chart(chart, f"A{chart_row}")

    out = output_path or workbook_path
    wb.save(out)
    return f"Added {chart_type} chart to sheet '{sheet_name}' → {out}"


# ── format ───────────────────────────────────────────────────────────
def apply_format(
    workbook_path: str,
    sheet_name: str,
    config_file: str,
    output_path: str = "",
) -> str:
    """Apply formatting to a workbook based on a JSON config.

    Config format:
    {
      "cells": {
        "A1": {"bold": true, "color": "FF0000", "fill": "FFFF00", "align": "center"},
        "B2:D10": {"number_format": "#,##0.00"}
      },
      "auto_width": true,
      "freeze": "A2"
    }
    """
    with open(config_file, "r", encoding="utf-8") as f:
        config = json.load(f)

    wb = openpyxl.load_workbook(workbook_path)
    ws = wb[sheet_name]

    # Parse cell ranges
    for range_str, style in config.get("cells", {}).items():
        cells = ws[range_str] if ":" in range_str else [ws[range_str]]
        if not isinstance(cells, (list, tuple)):
            cells = [cells]

        font_kwargs = {}
        if "bold" in style:
            font_kwargs["bold"] = style["bold"]
        if "color" in style:
            font_kwargs["color"] = style["color"]
        if "size" in style:
            font_kwargs["size"] = style["size"]
        font = Font(**font_kwargs) if font_kwargs else None

        fill = None
        if "fill" in style:
            fill = PatternFill(start_color=style["fill"], end_color=style["fill"],
                               fill_type="solid")

        align = None
        if "align" in style:
            align = Alignment(horizontal=style["align"], vertical="center")

        border = None
        if "border" in style:
            b = style["border"]
            side = Side(style=b.get("style", "thin"), color=b.get("color", "000000"))
            border = Border(left=side, right=side, top=side, bottom=side)

        for row in cells:
            for cell in row if hasattr(row, "__iter__") else [row]:
                if font:
                    cell.font = font
                if fill:
                    cell.fill = fill
                if align:
                    cell.alignment = align
                if border:
                    cell.border = border
                if "number_format" in style:
                    cell.number_format = style["number_format"]

    # Auto-width
    if config.get("auto_width"):
        _auto_width(ws)

    # Freeze panes
    if config.get("freeze"):
        ws.freeze_panes = config["freeze"]

    out = output_path or workbook_path
    wb.save(out)
    return f"Applied formatting to sheet '{sheet_name}' → {out}"


# ── info ─────────────────────────────────────────────────────────────
def workbook_info(workbook_path: str) -> str:
    """Show sheets, column names, and row counts for a workbook."""
    wb = openpyxl.load_workbook(workbook_path, read_only=True)
    lines: list[str] = [f"Workbook: {workbook_path}", f"Sheets: {len(wb.sheetnames)}"]
    for name in wb.sheetnames:
        ws = wb[name]
        cols = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1), [])]
        row_count = ws.max_row
        if row_count and row_count > 1:
            row_count -= 1  # exclude header
        else:
            row_count = 0
        lines.append(f"  [{name}] {len(cols)} cols, {row_count} data rows")
        lines.append(f"    Columns: {', '.join(str(c) for c in cols if c)}")
    wb.close()
    return "\n".join(lines)


# ── CLI ──────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="XLSX Creator Skill")
    sub = parser.add_subparsers(dest="action", required=True)

    # create
    p_create = sub.add_parser("create", help="Create a new workbook from data")
    p_create.add_argument("--source", required=True, help="CSV or JSON data file")
    p_create.add_argument("--output", required=True, help="Output .xlsx path")
    p_create.add_argument("--sheet-name", default="Sheet1", help="Worksheet name")
    p_create.add_argument("--no-auto-width", action="store_true", help="Skip auto-width")
    p_create.add_argument("--no-header-style", action="store_true", help="Skip header styling")

    # add-chart
    p_chart = sub.add_parser("add-chart", help="Add a chart to a workbook")
    p_chart.add_argument("--workbook", required=True, help="Path to .xlsx file")
    p_chart.add_argument("--sheet", default="Sheet1", help="Sheet name")
    p_chart.add_argument("--type", required=True, choices=["bar", "line", "pie", "scatter"])
    p_chart.add_argument("--data-start", required=True, help="Top-left cell (e.g. A1)")
    p_chart.add_argument("--data-end", required=True, help="Bottom-right cell (e.g. D10)")
    p_chart.add_argument("--categories-col", type=int, default=1)
    p_chart.add_argument("--output", default="", help="Output path (default: overwrite)")
    p_chart.add_argument("--title", default="", help="Chart title")
    p_chart.add_argument("--x-axis", default="", help="X-axis label")
    p_chart.add_argument("--y-axis", default="", help="Y-axis label")

    # format
    p_fmt = sub.add_parser("format", help="Apply formatting from JSON config")
    p_fmt.add_argument("--workbook", required=True, help="Path to .xlsx file")
    p_fmt.add_argument("--sheet", default="Sheet1", help="Sheet name")
    p_fmt.add_argument("--config", required=True, help="JSON config file")
    p_fmt.add_argument("--output", default="", help="Output path (default: overwrite)")

    # info
    p_info = sub.add_parser("info", help="Show workbook structure")
    p_info.add_argument("--workbook", required=True, help="Path to .xlsx file")

    args = parser.parse_args()

    try:
        if args.action == "create":
            print(create_workbook(
                args.source, args.output, args.sheet_name,
                auto_width=not args.no_auto_width,
                header_style=not args.no_header_style,
            ))
        elif args.action == "add-chart":
            print(add_chart(
                args.workbook, args.sheet, args.type,
                args.data_start, args.data_end,
                categories_col=args.categories_col,
                output_path=args.output,
                title=args.title, x_axis=args.x_axis, y_axis=args.y_axis,
            ))
        elif args.action == "format":
            print(apply_format(
                args.workbook, args.sheet, args.config, args.output,
            ))
        elif args.action == "info":
            print(workbook_info(args.workbook))
    except Exception as e:
        print(f"Error: {e}")
